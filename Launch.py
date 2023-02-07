import openpyxl
import os

# Creation of the file
destination_path = os.path.expanduser("~/My Documents/Commission Files/Commission Automation.xlsx")
commission_file = openpyxl.Workbook()

# Open Disputes data transfer
open_disputes = commission_file.active
open_disputes.title = "Open Disputes"
path_open_disputes = os.path.expanduser("~/My Documents/Commission Files/"
                                        "Open Disputes 101922 round 1.xlsx")
file_open_disputes = openpyxl.load_workbook(path_open_disputes)
open_disputes_master = commission_file["Open Disputes"]
# Copy and paste the data
sheet_open_disputes = file_open_disputes.active
max_row = sheet_open_disputes.max_row
for row in sheet_open_disputes.iter_rows(min_row=1, max_row=max_row, max_col=18, values_only=True):
    open_disputes_master.append(row)
# Adjusting the width of the cells
for col in open_disputes_master.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column letter
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    open_disputes_master.column_dimensions[column].width = adjusted_width

# Closing open disputes individual file
file_open_disputes.close()

# save the workbook
commission_file.save(destination_path)
